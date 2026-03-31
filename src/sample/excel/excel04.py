# xlsheet_read_to_dict.py
import openpyxl

# エクセルファイルの取り込み
wb = openpyxl.load_workbook("torihiki.xlsx")
ws = wb["Sheet1"]

# 2行目（列名のセル）
header_cells = ws[2]

# 3行目以降（データ）
student_list = []
for row in ws.iter_rows(min_row=3):
    row_dic = {}
    # セルの値を「key-value」で登録
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    student_list.append(row_dic)

print(student_list)