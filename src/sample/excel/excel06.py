import openpyxl

# 　コピーしたいファイル
inwb = openpyxl.load_workbook(r"d:\\_tmp\\test1.xlsx")

# 　ペーストしたいファイル
outwb = openpyxl.load_workbook(r"d:\\_tmp\\test2.xlsx")

# シートを選択（activeは現在開いているシート）
inws = inwb.active
outws = outwb.active

# 　シート名が決まっている場合は以下のようにシート名を指定する。
# 　outws = outwb["シート名"]


# 　最終行を取得。（一番下が何行まであるか）rowは行。1行目、2行目、3行目・・・
rows = inws.max_row

# 　最終列を取得。（一番横が何列まであるか）columnは列。A列、B列、C列・・・
columns = inws.max_column


# 　一列（横）をコピーして新しいところへ貼り付けるための関数
def input_excel(row_num):
    # コピーしたデータを入れるための空のリスト
    in_list = []

    # どこからどこまでの列をコピーしたいのか設定する。
    start_col = "A"
    goal_col = "CH"

    # 　A1:CH1　で数字の部分がrow_numで関数の引数に設定しているため最終行まで変更されていく。
    ranges = start_col + str(row_num) + ":" + goal_col + str(row_num)

    # 　in_listにコピーしたい行を追加していく。
    for i in inws[ranges]:
        for a in i:
            in_list.append(a.value)

    # 　in_listに追加されたデータをコピペしたいシートへ貼り付ける。
    outws.append(in_list)

    # 　コピペしたら保存する。
    outwb.save(r"d:\\_tmp\\test2.xlsx")


# 　コピー元のデータが入っている最終行までforで繰り返して、row_numへ数字を渡す。
for i in range(rows):
    i += 1
    print(i)
    input_excel(i)

# 　ペースト元の一番上の行が空白になるので削除する。
outws.delete_rows(1)
# 　保存する。
outwb.save(r"d:\\_tmp\\test2.xlsx")

# 閉じる
inwb.close()
outwb.close()