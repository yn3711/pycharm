import openpyxl

workbook = openpyxl.load_workbook('torihiki.xlsx')
sheet = workbook["Sheet1"]

suppliers = []

for i in range(3,9):
    cell_value = sheet.cell(row=i, column=2).value

    if cell_value not in suppliers:
        suppliers.append(cell_value)

print(suppliers)

transaction_amounts = []

for i in suppliers:
    transaction_amount = 0
    for j in range(3,9):
        supplier = sheet.cell(row=j, column=2).value

        if supplier == i :
            transaction_amount += sheet.cell(row=j, column=6).value

    transaction_amounts.append(transaction_amount)

print(transaction_amounts)

new_sheet = workbook.create_sheet("集計")

print(workbook.worksheets)

new_sheet["A1"] = "取引先"
new_sheet["B1"] = "取引金額"

for i, j, k in zip(list(range(3,9)), suppliers, transaction_amounts):
    new_sheet.cell(row=i, column=1, value=j)
    new_sheet.cell(row=i, column=2, value=k)

print(list(new_sheet.values))

workbook.save("torihiki.xlsx")